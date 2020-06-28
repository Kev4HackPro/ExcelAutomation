from openpyxl import load_workbook

wb =load_workbook(filename='reviews-sample.xlsx', data_only=True)
sheet_names = wb.sheetnames
print(sheet_names)
sheet = wb.active
print(wb.active)
print(sheet.title)

data1 = sheet['A1']
print(data1.value)
data2 = sheet['B1']
print(data2.value)
data3 =sheet['F7']
print(data3.value)