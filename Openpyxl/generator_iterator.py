import openpyxl as xl

wb = xl.load_workbook(filename='reviews-sample.xlsx', data_only=True)
sheet = wb.active

for row in sheet.iter_rows(min_row=1, 
                           max_row=2, 
                           min_col=1, 
                           max_col=3,
                           values_only=True):
    print(row)

for column in sheet.iter_cols(min_row=1,
                              max_row=2,
                              min_col=1,
                              max_col=3,
                              values_only=True):
    print(column)                         

# Iterating through the whole dataset
for row in sheet.rows:
    print(row)

for columns in sheet.columns:
    print(columns) 

for heeaders in sheet.iter_rows(min_row=1,
                                max_row=1,
                                values_only=True):
    print(heeaders)
                             
