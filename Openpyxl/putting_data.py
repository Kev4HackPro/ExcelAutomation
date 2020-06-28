import json
from openpyxl import load_workbook

wb = load_workbook(filename='reviews-sample.xlsx')
sheet = wb.active

products = {}

for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        'parent': row[1],
        'title': row[2],
        'category':row[3]
    }
    products[product_id] = product

print(json.dumps(products))