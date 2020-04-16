import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('insurance.xlsx')
sheet = wb['insurance']
for row in range(2, sheet.max_row + 1):
    ref_cell = sheet.cell(row, 7)
    corrected_price = ref_cell.value * 1.05
    corrected_price_cell = sheet.cell(row, 8)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,min_row=2,max_row = 50,min_col=7,max_col=7)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'i9')

wb.save('insurance_corrected_charges.xlsx')
